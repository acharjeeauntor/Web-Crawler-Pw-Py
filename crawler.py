from playwright.sync_api import Page
from openpyxl import Workbook
import os

def test_save_book_details(page:Page):
    page.goto("/")
    book_catagory_links = page.locator(".nav li ul li a")
    total_book_catagory_links = book_catagory_links.count()
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Book Name'
    ws['B1'] = 'Book Price'
    ws['C1'] = 'Book Availability'
    ws['D1'] = 'Book image Link'
    for index in range(total_book_catagory_links):
        book_catagory_link = book_catagory_links.nth(index)
        book_catagory_link.click()
        books_card = page.locator("ol li")
        total_books_count = books_card.count()
        for book in range(total_books_count):
            book_card = books_card.nth(book)
            image_link = book_card.locator(".image_container a").get_attribute('href')
            book_name = book_card.locator("h3 a").text_content()
            book_price = book_card.locator(".price_color").text_content()
            book_availability = book_card.locator("[class='instock availability']").text_content()
            ws.append([book_name,book_price,book_availability,image_link])

    file_directory = 'Book_List_EXCEL_Data'
    if not os.path.exists(file_directory):
        os.makedirs(file_directory)
        
    file_path = os.path.join(file_directory, 'book_data.xlsx')

    # Create the directory if it doesn't exist
    wb.save(file_path)        